"""
Importación de liquidaciones históricas desde Excel (.xlsx) o CSV.
Permite cargar lo que la contadora ya liquidó, mes a mes, para tener
toda la historia en un solo sistema.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.orm import Session
from datetime import date
import csv, io

import openpyxl

from ..database import get_db
from ..models import EmpleadoDB, LiquidacionDB, UserDB
from ..auth import get_current_user
from .activity_log import registrar

router = APIRouter(prefix="/importar", tags=["importar"])

# Orden y nombres de columnas de la plantilla
COLUMNAS = [
    "rut", "nombre", "anio", "mes", "dias_trabajados",
    "sueldo_mes", "gratificacion", "bonos", "horas_extras", "comisiones",
    "colacion", "movilizacion",
    "fondo_pensiones", "comision_afp", "afc_trabajador", "salud",
    "impuesto_unico", "otros_descuentos", "liquido_a_pagar",
]

FILA_EJEMPLO = [
    "16.630.410-5", "Alexandra Albornoz", 2026, 5, 30,
    602488, 175622, 100000, 0, 0,
    98813, 0,
    87811, 12733, 5269, 61468,
    7373, 0, 802269,
]


def _norm_rut(rut) -> str:
    """Normaliza RUT para comparar: sin puntos, guiones ni ceros a la izquierda."""
    r = "".join(c for c in str(rut).upper() if c.isalnum())
    return r.lstrip("0")


def _num(valor) -> float:
    """Convierte celdas tipo '$1.234.567', '1234567' o vacías a número."""
    if valor is None or valor == "":
        return 0
    if isinstance(valor, (int, float)):
        return float(valor)
    limpio = "".join(c for c in str(valor) if c.isdigit() or c in "-.,")
    # Formato chileno: puntos de miles → eliminarlos; coma decimal → punto
    limpio = limpio.replace(".", "").replace(",", ".")
    try:
        return float(limpio)
    except ValueError:
        return 0


@router.get("/plantilla")
def plantilla(_: UserDB = Depends(get_current_user)):
    """Descarga la plantilla Excel con encabezados y una fila de ejemplo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"
    ws.append(COLUMNAS)
    ws.append(FILA_EJEMPLO)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_liquidaciones.xlsx"'},
    )


def _leer_filas(archivo: UploadFile, contenido: bytes) -> list[dict]:
    """Lee xlsx o csv y devuelve lista de dicts con claves de COLUMNAS."""
    nombre = (archivo.filename or "").lower()

    if nombre.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active
        filas = list(ws.iter_rows(values_only=True))
    else:
        texto = contenido.decode("utf-8-sig", errors="replace")
        delim = ";" if texto.split("\n")[0].count(";") >= texto.split("\n")[0].count(",") else ","
        filas = list(csv.reader(io.StringIO(texto), delimiter=delim))

    if not filas:
        return []

    encabezados = [str(h or "").strip().lower() for h in filas[0]]
    resultado = []
    for fila in filas[1:]:
        if not fila or all(c in (None, "") for c in fila):
            continue
        d = dict(zip(encabezados, fila))
        resultado.append(d)
    return resultado


@router.post("/liquidaciones")
async def importar_liquidaciones(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: UserDB = Depends(get_current_user),
):
    contenido = await archivo.read()
    try:
        filas = _leer_filas(archivo, contenido)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo: {e}")

    if not filas:
        raise HTTPException(400, "El archivo no tiene filas de datos")

    # Mapa RUT normalizado → empleado existente
    empleados = {_norm_rut(e.rut): e for e in db.query(EmpleadoDB).all()}

    creadas = actualizadas = empleados_creados = 0
    errores = []

    for i, fila in enumerate(filas, start=2):  # fila 1 = encabezados
        try:
            rut = str(fila.get("rut") or "").strip()
            anio = int(_num(fila.get("anio")))
            mes  = int(_num(fila.get("mes")))
            if not rut or not anio or not (1 <= mes <= 12):
                errores.append({"fila": i, "error": "Falta rut, año o mes válido"})
                continue

            # La fila de ejemplo de la plantilla se ignora si quedó sin editar
            if rut == FILA_EJEMPLO[0] and str(fila.get("nombre", "")).strip() == FILA_EJEMPLO[1] \
               and anio == FILA_EJEMPLO[2] and mes == FILA_EJEMPLO[3] and len(filas) > 1:
                continue

            emp = empleados.get(_norm_rut(rut))
            if not emp:
                nombre = str(fila.get("nombre") or "").strip()
                if not nombre:
                    errores.append({"fila": i, "error": f"RUT {rut} no existe y no trae nombre para crearlo"})
                    continue
                emp = EmpleadoDB(
                    nombre=nombre,
                    rut=rut,
                    fecha_inicio=date(anio, mes, 1),
                    sueldo_base=_num(fila.get("sueldo_mes")),
                    activo=True,
                )
                db.add(emp)
                db.flush()
                empleados[_norm_rut(rut)] = emp
                empleados_creados += 1

            dias = int(_num(fila.get("dias_trabajados")) or 30)

            haberes_imp = (
                _num(fila.get("sueldo_mes")) + _num(fila.get("gratificacion"))
                + _num(fila.get("bonos")) + _num(fila.get("horas_extras"))
                + _num(fila.get("comisiones"))
            )
            haberes_no_imp = _num(fila.get("colacion")) + _num(fila.get("movilizacion"))
            total_haberes  = haberes_imp + haberes_no_imp
            total_desc = (
                _num(fila.get("fondo_pensiones")) + _num(fila.get("comision_afp"))
                + _num(fila.get("afc_trabajador")) + _num(fila.get("salud"))
                + _num(fila.get("impuesto_unico")) + _num(fila.get("otros_descuentos"))
            )
            liquido = _num(fila.get("liquido_a_pagar")) or (total_haberes - total_desc)

            resultado = {
                "sueldo_mes":            _num(fila.get("sueldo_mes")),
                "gratificacion":         _num(fila.get("gratificacion")),
                "bonos_fijos":           _num(fila.get("bonos")),
                "horas_extras":          _num(fila.get("horas_extras")),
                "comisiones":            _num(fila.get("comisiones")),
                "colacion":              _num(fila.get("colacion")),
                "movilizacion":          _num(fila.get("movilizacion")),
                "haberes_imponibles":    haberes_imp,
                "haberes_no_imponibles": haberes_no_imp,
                "total_haberes":         total_haberes,
                "fondo_pensiones":       _num(fila.get("fondo_pensiones")),
                "comision_afp":          _num(fila.get("comision_afp")),
                "afc_trabajador":        _num(fila.get("afc_trabajador")),
                "salud":                 _num(fila.get("salud")),
                "impuesto_unico":        _num(fila.get("impuesto_unico")),
                "otros_descuentos":      _num(fila.get("otros_descuentos")),
                "total_descuentos":      total_desc,
                "base_tributable":       max(0, haberes_imp - total_desc + _num(fila.get("impuesto_unico"))),
                "liquido_a_pagar":       liquido,
                "dias_trabajados":       dias,
                "importado":             True,
                "fuente":                "histórico contadora",
            }

            existente = db.query(LiquidacionDB).filter(
                LiquidacionDB.empleado_id == emp.id,
                LiquidacionDB.mes == mes,
                LiquidacionDB.anio == anio,
            ).first()

            if existente:
                existente.resultado = resultado
                existente.dias_trabajados = dias
                actualizadas += 1
            else:
                db.add(LiquidacionDB(
                    empleado_id=emp.id, mes=mes, anio=anio,
                    dias_trabajados=dias, resultado=resultado,
                ))
                creadas += 1

        except Exception as e:
            errores.append({"fila": i, "error": str(e)})

    db.commit()
    registrar(db, user, "importar", "liquidacion", None,
              f"{creadas} creadas, {actualizadas} actualizadas, {empleados_creados} empleados nuevos")

    return {
        "ok": True,
        "creadas": creadas,
        "actualizadas": actualizadas,
        "empleados_creados": empleados_creados,
        "errores": errores,
    }
